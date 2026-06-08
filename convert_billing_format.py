"""
convert_billing_format.py

Reads Import_Sample.xlsx (IOP multi-row format) and writes
Import_Converted.xlsx in the flat target format.

Data cleaning applied:
  1. Year added to bare M/D dates/ranges    4/23-4/25   ->  4/23/2026-4/25/2026
  2. Consecutive dates merged               4/27, 4/28, 4/29  ->  4/27/2026-4/29/2026
  3. Non-consecutive stay separate          3/26, 3/30-3/31   ->  row1: 3/26/2026
                                                                    row2: 3/30/2026-3/31/2026
  4. H0038 units format                     3/19-2; 3/25-3    ->  row1: 3/19/2026-2
                                                                    row2: 3/25/2026-3
  5. Noise removed                          Approved, PENDING, ?????,
                                            auth-unit numbers (35, 10), header strings
  6. Parenthetical stripped from dates      2/27 (2)          ->  2/27/2026

Source column mapping (Import_Sample.xlsx col -> service code):
    G(7)  OP GROUP DATES          -> H0005     $220.65  (dates)
    H(8)  OP INDIVIDUAL SESSIONS  -> H0004     $124.24  (dates)
    I(9)  # IOP AUTH UNITS        -> H0015     $194.23  (dates only; numbers ignored)
    J(10) IOP GROUP DATES         -> H0015     $194.23  (dates)
    K(11) # PHP AUTH UNITS        -> H2036      $32.38  (dates only; numbers ignored)
    L(12) PHP GROUP DATES         -> H2036      $32.38  (dates)
    M(13) URINALYSIS              -> UA        $326.27  (dates)
    N(14) Peer Group Units        -> MDRN       $34.37  (dates)
    O(15) PEER GROUP DATES        -> H0024      $10.22  (dates)
    Q(17) PEER IND. UNITS         -> H0038      $18.77  (units M/D-N)
"""

import re
import warnings
from datetime import datetime, date, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

INPUT_FILE  = 'Import_Sample.xlsx'
OUTPUT_FILE = 'Import_Converted.xlsx'

SOURCE_TO_SERVICE = {
    7:  'H0005',
    8:  'H0004',
    9:  'H0015',   # # IOP AUTH UNITS — dates only
    10: 'H0015',
    11: 'H2036',   # # PHP AUTH UNITS — dates only
    12: 'H2036',
    13: 'UA',
    14: 'MDRN',
    15: 'H0024',
    16: 'H0038',   # Peer Ind. Units — dates only (Q is units M/D-N)
    17: 'H0038',   # units format: M/D-N
}

# Target column order matches customer billing data.xlsx rows 4-8
TARGET_HEADERS = [
    'First Name', 'Last Name', 'DOB', 'Status', 'ID', 'F ID',
    'H0005', 'H0004', 'H0015', 'H2036', 'UA', 'MDRN', 'H0024', 'H0038',
]
TARGET_SVC_ORDER = [
    'H0005', 'H0004', 'H0015', 'H2036', 'UA', 'MDRN', 'H0024', 'H0038',
]


# ---------------------------------------------------------------------------
# Year inference
# ---------------------------------------------------------------------------

def _infer_year(group_rows, ws, fallback=2026):
    """Return the year from the first datetime found in service cols 7-17."""
    for r in group_rows:
        for c in range(7, 18):
            val = ws.cell(r, c).value
            if isinstance(val, datetime):
                return val.year
    return fallback


# ---------------------------------------------------------------------------
# Low-level date helpers
# ---------------------------------------------------------------------------

def _parse_single_date(token, year):
    """
    Parse 'M/D', 'M/D/YYYY', or 'M/D/YY' (2-digit year -> 20YY).
    Returns a date object or None.
    """
    t = token.strip()
    # M/D/YYYY
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', t)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    # M/D/YY -> treat as 20YY
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2})$', t)
    if m:
        try:
            return date(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    # M/D
    m = re.match(r'^(\d{1,2})/(\d{1,2})$', t)
    if m:
        try:
            return date(year, int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def _expand_range(start, end):
    if end < start:                          # year wrap (e.g. 12/30-1/3)
        end = end.replace(year=end.year + 1)
    result, cur = [], start
    while cur <= end:
        result.append(cur)
        cur += timedelta(days=1)
    return result


def _dates_to_range_strings(sorted_dates):
    """
    Merge consecutive dates and format.
    Returns list of 'M/D/YYYY' or 'M/D/YYYY-M/D/YYYY' strings.
    """
    if not sorted_dates:
        return []
    dates = sorted(set(sorted_dates))
    groups, s, e = [], dates[0], dates[0]
    for d in dates[1:]:
        if (d - e).days == 1:
            e = d
        else:
            groups.append((s, e))
            s = e = d
    groups.append((s, e))

    out = []
    for s, e in groups:
        if s == e:
            out.append(f"{s.month}/{s.day}/{s.year}")
        else:
            out.append(f"{s.month}/{s.day}/{s.year}-{e.month}/{e.day}/{e.year}")
    return out


# ---------------------------------------------------------------------------
# Date-column cell parser
# ---------------------------------------------------------------------------

def _parse_one_date_token(token, year):
    """
    Parse a single date token (no commas/semicolons inside).
    Handles: M/D, M/D/YYYY, M/D-M/D, M/D/YYYY-M/D/YYYY, M/D-M/D/YY, M/D (N).
    Returns list of date objects.
    """
    t = token.strip()
    if not t or '/' not in t:
        return []

    # Strip trailing parenthetical e.g. '2/27 (2)' -> '2/27'
    t = re.sub(r'\s*\(\d+\)\s*$', '', t).strip()
    if not t:
        return []

    # Range M/D/YYYY-M/D/YYYY  (with optional spaces around the dash)
    m = re.match(r'^(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})$', t)
    if m:
        sd = _parse_single_date(m.group(1), year)
        ed = _parse_single_date(m.group(2), year)
        if sd and ed:
            return _expand_range(sd, ed)

    # Range M/D-M/D/YY  e.g. '3/1-3/3/26'
    m = re.match(r'^(\d{1,2}/\d{1,2})\s*-\s*(\d{1,2}/\d{1,2}/\d{2})$', t)
    if m:
        sd = _parse_single_date(m.group(1), year)
        ed = _parse_single_date(m.group(2), year)
        if sd and ed:
            return _expand_range(sd, ed)

    # Range M/D-M/D
    m = re.match(r'^(\d{1,2}/\d{1,2})\s*-\s*(\d{1,2}/\d{1,2})$', t)
    if m:
        sd = _parse_single_date(m.group(1), year)
        ed = _parse_single_date(m.group(2), year)
        if sd and ed:
            return _expand_range(sd, ed)

    # Single date (M/D, M/D/YYYY, M/D/YY)
    d = _parse_single_date(t, year)
    if d:
        return [d]

    return []


def _parse_date_cell(raw_val, year):
    """
    Parse a full date-column cell value into a list of date objects.
    Handles comma-separated entries.  Silently drops noise.
    """
    if raw_val is None:
        return []
    if isinstance(raw_val, datetime):
        return [raw_val.date()]
    if isinstance(raw_val, (int, float)):
        return []   # auth unit count only (no date) — skip

    s = str(raw_val).strip()
    if not s or '/' not in s:
        return []   # 'Approved', 'PENDING', '?????', column header strings, etc.

    results = []
    for token in s.split(','):
        results.extend(_parse_one_date_token(token, year))
    return results


# ---------------------------------------------------------------------------
# H0038 cell parser
# ---------------------------------------------------------------------------

def _parse_h0038_cell(raw_val, year):
    """
    Parse a PEER INDIVIDUAL SESSION cell.
    Accepts M/D-N or M/D (N) entries, semicolon- or comma-separated.
    Returns list of strings 'M/D/YYYY-N'.
    """
    if raw_val is None:
        return []
    if isinstance(raw_val, datetime):
        return [f"{raw_val.month}/{raw_val.day}/{raw_val.year}-1"]
    if isinstance(raw_val, (int, float)):
        return []

    s = str(raw_val).strip()
    if not s or '/' not in s:
        return []

    entries = []
    for token in re.split(r'[;,]', s):
        token = token.strip()
        if not token:
            continue

        # M/D-N  (units with dash, optional spaces)
        m = re.match(r'^(\d{1,2}/\d{1,2})\s*-\s*(\d+)\s*$', token)
        if m:
            d = _parse_single_date(m.group(1), year)
            if d:
                entries.append(f"{d.month}/{d.day}/{d.year}-{m.group(2)}")
            continue

        # M/D/YYYY-N
        m = re.match(r'^(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d+)\s*$', token)
        if m:
            d = _parse_single_date(m.group(1), year)
            if d:
                entries.append(f"{d.month}/{d.day}/{d.year}-{m.group(2)}")
            continue

        # M/D (N)  — parenthetical units
        m = re.match(r'^(\d{1,2}/\d{1,2})\s*\((\d+)\)\s*$', token)
        if m:
            d = _parse_single_date(m.group(1), year)
            if d:
                entries.append(f"{d.month}/{d.day}/{d.year}-{m.group(2)}")
            continue

        # Bare M/D with no units — treat as 1 unit
        d = _parse_single_date(token, year)
        if d:
            entries.append(f"{d.month}/{d.day}/{d.year}-1")

    return entries


# ---------------------------------------------------------------------------
# Customer metadata extraction
# ---------------------------------------------------------------------------

def _extract_meta(start_row, group_rows, ws):
    """
    Pull name, DOB, MD ID, and F-diagnosis code from the multi-row customer group.
    Header row: col B=last_name, col C=first_name, col D=F-diagnosis (e.g. F11.20)
    Sub-rows:   col B = MD ID | DOB | insurance# | address  (any order)
    """
    last_name  = str(ws.cell(start_row, 2).value or '').strip()
    first_name = str(ws.cell(start_row, 3).value or '').strip()
    # F ID is the diagnosis code (e.g. F11.20) from col D of the header row
    f_id       = str(ws.cell(start_row, 4).value or '').strip() or None
    md_id = dob = None

    for r in group_rows[1:]:
        b = ws.cell(r, 2).value
        if b is None:
            continue
        if isinstance(b, datetime):
            if dob is None:
                dob = f"{b.month}/{b.day}/{b.year}"
        elif isinstance(b, str):
            bv = b.strip()
            if re.match(r'^MD\d+$', bv, re.IGNORECASE) and md_id is None:
                md_id = bv.upper()

    return last_name, first_name, dob, md_id, f_id


# ---------------------------------------------------------------------------
# Per-customer service data collector
# ---------------------------------------------------------------------------

def _collect_services(group_rows, ws, year):
    """
    Parse and clean all service data for a customer group.
    Returns {svc_name: [entry1, entry2, ...]} with consecutive dates merged.
    """
    raw = {col: [] for col in SOURCE_TO_SERVICE}
    for r in group_rows:
        for col_idx in SOURCE_TO_SERVICE:
            v = ws.cell(r, col_idx).value
            if v is not None:
                raw[col_idx].append(v)

    result = {}
    for col_idx, svc_name in SOURCE_TO_SERVICE.items():
        vals = raw[col_idx]
        if not vals:
            continue

        if col_idx == 17:
            entries = []
            for v in vals:
                entries.extend(_parse_h0038_cell(v, year))
            if entries:
                result.setdefault(svc_name, []).extend(entries)
        else:
            all_dates = []
            for v in vals:
                all_dates.extend(_parse_date_cell(v, year))
            if all_dates:
                result.setdefault(svc_name, []).extend(
                    _dates_to_range_strings(all_dates)
                )

    return result


# ---------------------------------------------------------------------------
# Output row builder
# ---------------------------------------------------------------------------

def _build_row(customer_info, svc_dict):
    """customer_info = (first, last, dob, md_id, f_id) or None for continuation rows."""
    if customer_info:
        first, last, dob, md_id, f_id = customer_info
        row = [first, last, dob, 'Active', md_id, f_id]
    else:
        row = [None] * 6
    for svc in TARGET_SVC_ORDER:
        row.append(svc_dict.get(svc))
    return row


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb_in = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    # Try to read the global date from row 1 for year fallback
    ws0 = wb_in.active
    global_date_val = ws0.cell(1, 2).value
    fallback_year = (global_date_val.year
                     if isinstance(global_date_val, datetime)
                     else 2026)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Converted'

    # Header row with styling
    ws_out.append(TARGET_HEADERS)
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF')
    for cell in ws_out[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    total_customers = 0
    total_rows      = 0

    for sheet_name in wb_in.sheetnames:
        ws = wb_in[sheet_name]

        # Customer header rows = col A has a positive integer
        customer_starts = []
        for r in range(1, ws.max_row + 1):
            a_val = ws.cell(r, 1).value
            if a_val is None:
                continue
            try:
                if int(float(str(a_val))) > 0:
                    customer_starts.append(r)
            except (ValueError, TypeError):
                pass

        for idx, start_row in enumerate(customer_starts):
            end_row    = (customer_starts[idx + 1] - 1
                          if idx + 1 < len(customer_starts) else ws.max_row)
            group_rows = list(range(start_row, end_row + 1))

            last_name, first_name, dob, md_id, f_id = _extract_meta(
                start_row, group_rows, ws
            )
            if not last_name:
                continue

            year     = _infer_year(group_rows, ws, fallback=fallback_year)
            services = _collect_services(group_rows, ws, year)
            info     = (first_name, last_name, dob, md_id, f_id)

            max_entries = max((len(v) for v in services.values()), default=0)

            if max_entries == 0:
                ws_out.append(_build_row(info, {}))
                total_rows += 1
            else:
                for i in range(max_entries):
                    svc_slice = {k: v[i] for k, v in services.items() if i < len(v)}
                    ws_out.append(_build_row(info if i == 0 else None, svc_slice))
                    total_rows += 1

            total_customers += 1

    # Auto-size columns
    for col in ws_out.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    ws_out.freeze_panes = 'A2'
    wb_out.save(OUTPUT_FILE)

    print(f'Done.')
    print(f'  Customers processed : {total_customers}')
    print(f'  Output rows written : {total_rows}')
    print(f'  Output file         : {OUTPUT_FILE}')


if __name__ == '__main__':
    main()
