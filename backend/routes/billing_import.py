from fastapi import APIRouter, UploadFile, File, HTTPException, Query, status, Depends, Body
from pydantic import BaseModel, Field
import openpyxl
import re
import random
import string
import warnings
import json
import io
import requests
from urllib.parse import unquote
from datetime import datetime, date, timedelta, timezone
from typing import Optional, List, Tuple, Dict, Any

from db.database import get_db_connection
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import os

router = APIRouter()
security = HTTPBearer()


# ---------------------------------------------------------------------------
# Service column mapping  (1-based col index -> (service_code, col_type))
# col_type 'date'  -> dates collected across all rows, consecutive merged
# col_type 'units' -> cell format "M/D-N"; N = unit count (no date merging)
# Auth columns (9, 11): dates imported; plain numbers (e.g. 35, 10) are ignored.
#
# Rates are NOT stored here — they're read live from the `services` table
# (see _get_service_rates) so an admin-edited rate takes effect on the next
# import without a code change. If a mapped service has no row in `services`,
# the import fails fast with a clear error rather than guessing a rate.
# ---------------------------------------------------------------------------
SERVICE_COLS = {
    7:  ('H0005','date'),   # G: OP GROUP DATES
    8:  ('H0004', 'date'),   # H: OP INDIVIDUAL SESSIONS
    10: ('H0015', 'date'),   # J: IOP GROUP DATES
    12: ('H2036', 'date'),   # L: PHP GROUP DATES
    13: ('UA',    'date'),   # M: URINALYSIS
    15: ('H0024', 'date'),   # O: PEER GROUP DATES
    16: ('H0038', 'units'),  # P: PEER INDIVIDUAL UNITS
    # I(9)=IOP AUTH UNITS, K(11)=PHP AUTH UNITS, N(14)=Peer Group Units,
    # Q(17)=Peer Individual Session — not mapped, data ignored
}


def _build_service_col_types() -> Dict[str, str]:
    """service_code -> col_type ('date' | 'units'), for the fix-invalid flow."""
    lookup: Dict[str, str] = {}
    for _, (code, col_type) in SERVICE_COLS.items():
        if code not in lookup or col_type == 'units':
            lookup[code] = col_type
    return lookup


SERVICE_COL_TYPES: Dict[str, str] = _build_service_col_types()


def _get_service_rates(cur) -> Dict[str, float]:
    """Current rate_per_day per service, read live from the `services` table —
    the single source of truth for billing rates (edited on the Services page).
    """
    rates: Dict[str, float] = {}
    cur.execute('SELECT name, rate_per_day FROM services')
    for row in cur.fetchall():
        name = row['name'] if isinstance(row, dict) else row[0]
        rate = row['rate_per_day'] if isinstance(row, dict) else row[1]
        if name is not None and rate is not None:
            rates[name] = float(rate)
    return rates


def _require_rates_for_mapped_services(rates: Dict[str, float]) -> None:
    """Fail fast with a clear error if any SERVICE_COLS-mapped service has no
    rate configured in the `services` table, instead of silently billing $0.
    """
    mapped_codes = {code for _, (code, _col_type) in SERVICE_COLS.items()}
    missing = sorted(code for code in mapped_codes if code not in rates)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Missing rate_per_day in the Services table for: "
                f"{', '.join(missing)}. Add/fix these on the Services page, then retry the import."
            ),
        )


ERR_MONTH_ORDER = (
    'Start month is after end month — add year (e.g. 3/15/2025-2/20/2026)'
)


# ---------------------------------------------------------------------------
# Date parsing helpers  (shared with convert_billing_format.py logic)
# ---------------------------------------------------------------------------

def _infer_year(group_rows: List[int], ws) -> int:
    """Return the year from the first datetime found in service cols 7-17."""
    for r in group_rows:
        for c in range(7, 18):
            val = ws.cell(r, c).value
            if isinstance(val, datetime):
                return val.year
    header_date = ws.cell(1, 2).value
    if isinstance(header_date, datetime):
        return header_date.year
    return datetime.utcnow().year


def _parse_single_date(token: str, year: int) -> Optional[date]:
    """Parse 'M/D', 'M/D/YYYY', or 'M/D/YY' (2-digit -> 20YY)."""
    t = token.strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', t)
    if m:
        try:
            mo, dy, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if mo < 1 or mo > 12:
                return None
            return date(yr, mo, dy)
        except ValueError:
            return None
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2})$', t)
    if m:
        try:
            mo, dy = int(m.group(1)), int(m.group(2))
            if mo < 1 or mo > 12:
                return None
            return date(2000 + int(m.group(3)), mo, dy)
        except ValueError:
            return None
    m = re.match(r'^(\d{1,2})/(\d{1,2})$', t)
    if m:
        try:
            mo, dy = int(m.group(1)), int(m.group(2))
            if mo < 1 or mo > 12:
                return None
            return date(year, mo, dy)
        except ValueError:
            return None
    return None


def _expand_dates(start: date, end: date) -> List[date]:
    if end < start:
        end = end.replace(year=end.year + 1)
    result, cur = [], start
    while cur <= end:
        result.append(cur)
        cur += timedelta(days=1)
    return result


def _bare_range_month_order(start_part: str, end_part: str) -> bool:
    """True when M/D-M/D has start month > end month (year required)."""
    sm = re.match(r'^(\d{1,2})/(\d{1,2})$', start_part.strip())
    em = re.match(r'^(\d{1,2})/(\d{1,2})$', end_part.strip())
    if not sm or not em:
        return False
    return int(sm.group(1)) > int(em.group(1))


def _parse_one_date_token(token: str, year: int) -> Tuple[List[date], Optional[str]]:
    """
    Parse a single date token (no commas/semicolons).
    Returns (dates, error_message). error_message set when year must be added.
    """
    t = token.strip()
    if not t or '/' not in t:
        return [], None
    # Strip trailing parenthetical e.g. '2/27 (2)' -> '2/27'
    t = re.sub(r'\s*\(\d+\)\s*$', '', t).strip()
    if not t:
        return [], None

    # Range M/D/YYYY-M/D/YYYY
    m = re.match(r'^(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})$', t)
    if m:
        sd, ed = _parse_single_date(m.group(1), year), _parse_single_date(m.group(2), year)
        if sd and ed:
            return _expand_dates(sd, ed), None
        return [], 'Invalid date format — expected M/D/YYYY-M/D/YYYY'

    # Range M/D-M/D/YY  e.g. '3/1-3/3/26'
    m = re.match(r'^(\d{1,2}/\d{1,2})\s*-\s*(\d{1,2}/\d{1,2}/\d{2})$', t)
    if m:
        sd, ed = _parse_single_date(m.group(1), year), _parse_single_date(m.group(2), year)
        if sd and ed:
            return _expand_dates(sd, ed), None
        return [], 'Invalid date format — expected M/D-M/D/YY'

    # Range M/D-M/D — flag when start month > end month (user must add year)
    m = re.match(r'^(\d{1,2}/\d{1,2})\s*-\s*(\d{1,2}/\d{1,2})$', t)
    if m:
        if _bare_range_month_order(m.group(1), m.group(2)):
            return [], ERR_MONTH_ORDER
        sd, ed = _parse_single_date(m.group(1), year), _parse_single_date(m.group(2), year)
        if sd and ed:
            return _expand_dates(sd, ed), None
        return [], ERR_MONTH_ORDER

    # Single date
    d = _parse_single_date(t, year)
    if d:
        return [d], None
    if '/' in t:
        return [], 'Invalid date format — expected M/D or M/D/YYYY'
    return [], None


def _parse_date_cell(raw_val, year: int) -> Tuple[List[date], Optional[str]]:
    """
    Parse a date-column cell value.
    Returns (dates, first_error). Silently drops noise (Approved, ?????, numbers, etc.).
    """
    if raw_val is None:
        return [], None
    if isinstance(raw_val, datetime):
        return [raw_val.date()], None
    if isinstance(raw_val, (int, float)):
        return [], None
    s = str(raw_val).strip()
    if not s or '/' not in s:
        return [], None
    results: List[date] = []
    first_error: Optional[str] = None
    for token in s.split(','):
        parsed, err = _parse_one_date_token(token, year)
        if parsed:
            results.extend(parsed)
        elif err and not first_error:
            first_error = err
    return results, first_error


def _dates_to_range_entries(all_dates: List[date], rate: float) -> List[Dict[str, Any]]:
    """
    Merge consecutive dates and return DB-ready entry dicts.
    Each merged range -> one entry: days = span, amount = days * rate.
    """
    if not all_dates:
        return []
    dates = sorted(set(all_dates))
    groups: List[Tuple[date, date]] = []
    s = e = dates[0]
    for d in dates[1:]:
        if (d - e).days == 1:
            e = d
        else:
            groups.append((s, e))
            s = e = d
    groups.append((s, e))

    entries = []
    for s, e in groups:
        days = (e - s).days + 1
        entries.append({
            'start_date':    s.strftime('%Y-%m-%d'),
            'end_date':      e.strftime('%Y-%m-%d'),
            'days':          days,
            'rate_per_day':  rate,
            'amount_billed': round(days * rate, 2),
        })
    return entries


def _parse_h0038_cell(raw_val, year: int, rate: float) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Parse a PEER INDIVIDUAL SESSION cell.
    Accepts M/D-N or M/D (N) entries, semicolon- or comma-separated.
    Returns DB-ready entry dicts; does NOT merge entries.
    """
    if raw_val is None:
        return [], None
    if isinstance(raw_val, datetime):
        dt_str = raw_val.strftime('%Y-%m-%d')
        return [{'start_date': dt_str, 'end_date': dt_str, 'days': 1, 'units': 1,
                 'rate_per_day': rate, 'amount_billed': round(rate, 2)}], None
    if isinstance(raw_val, (int, float)):
        return [], None
    s = str(raw_val).strip()
    if not s or '/' not in s:
        return [], None

    entries = []
    first_error: Optional[str] = None
    for token in re.split(r'[;,]', s):
        token = token.strip()
        if not token:
            continue
        d = units = None

        # M/D-N
        m = re.match(r'^(\d{1,2}/\d{1,2})\s*-\s*(\d+)\s*$', token)
        if m:
            d, units = _parse_single_date(m.group(1), year), int(m.group(2))
        # M/D/YYYY-N
        if not d:
            m = re.match(r'^(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d+)\s*$', token)
            if m:
                d, units = _parse_single_date(m.group(1), year), int(m.group(2))
        # M/D (N)
        if not d:
            m = re.match(r'^(\d{1,2}/\d{1,2})\s*\((\d+)\)\s*$', token)
            if m:
                d, units = _parse_single_date(m.group(1), year), int(m.group(2))
        # Bare date — 1 unit
        if not d:
            d = _parse_single_date(token, year)
            if d:
                units = 1

        if d and units:
            dt_str = d.strftime('%Y-%m-%d')
            entries.append({
                'start_date':    dt_str,
                'end_date':      dt_str,
                # H0038 is billed per unit, not per day — each entry covers one
                # calendar day but a variable number of units (15-min sessions).
                'days':          1,
                'units':         units,
                'rate_per_day':  rate,
                'amount_billed': round(units * rate, 2),
            })
        elif '/' in token and not first_error:
            _, err = _parse_one_date_token(token, year)
            first_error = err or 'Invalid unit format — expected M/D-N (e.g. 3/5-2)'
    if entries:
        return entries, None
    return [], first_error or 'Invalid unit format — expected M/D-N (e.g. 3/5-2)'


# ---------------------------------------------------------------------------
# Member block detection
# ---------------------------------------------------------------------------

def _find_member_blocks(ws) -> List[Tuple[int, List[int]]]:
    """
    Scan a worksheet and group rows into member blocks.

    A member block starts at any row where col A holds a positive integer
    (the spreadsheet's member number — NOT the application's customer/user
    identity, see _resolve_customer) and runs up to — but not including —
    the next such row, or the end of the sheet. Blank rows, and rows whose
    customer/billing data is offset from the member-number row by any number
    of rows, all stay inside the same block; only another valid member-number
    row (or the end of the sheet) closes it.

    Returns a list of (start_row, group_rows) tuples, one per member block,
    in sheet order.
    """
    member_number_rows = []
    for r in range(1, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        if a_val is None:
            continue
        try:
            if int(float(str(a_val))) > 0:
                member_number_rows.append(r)
        except (ValueError, TypeError):
            pass

    blocks = []
    for i, start_row in enumerate(member_number_rows):
        end_row = member_number_rows[i + 1] - 1 if i + 1 < len(member_number_rows) else ws.max_row
        blocks.append((start_row, list(range(start_row, end_row + 1))))
    return blocks


# ---------------------------------------------------------------------------
# Customer extraction helpers
# ---------------------------------------------------------------------------

def _extract_customer_info(start_row: int, group_rows: List[int], ws):
    """
    Search the ENTIRE member block for identifying fields — their row position
    is NOT guaranteed to be start_row (the member-number row). A member's
    name/diagnosis/assessment can appear several rows later, e.g. after blank
    rows, if the spreadsheet's layout varies between members.

    Name row (found anywhere in the block): col B=last_name, col C=first_name,
    col D=F-diagnosis (F ID), col F=assessment. The first row in the block
    where both B and C hold non-empty, non-MD-id string values is taken as
    the name row — for the common case this is start_row itself, preserving
    existing behavior for spreadsheets that already work.

    Sub-rows (anywhere in the block): col B = MD ID ('MD...') | DOB (datetime).
    """
    last_name = first_name = ''
    f_id: Optional[str] = None
    assessment_date: Optional[str] = None
    md_id = dob = None

    for r in group_rows:
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value

        if not last_name and isinstance(b, str) and isinstance(c, str):
            b_val = b.strip()
            c_val = c.strip()
            if b_val and c_val and not re.match(r'^MD\d+$', b_val, re.IGNORECASE):
                last_name = b_val.upper()
                first_name = c_val.upper()
                d_val = ws.cell(r, 4).value
                f_id = str(d_val).strip() if d_val not in (None, '') else None
                assess_val = ws.cell(r, 6).value
                if isinstance(assess_val, datetime):
                    assessment_date = assess_val.strftime('%Y-%m-%d')
                continue  # this row's B/C already consumed as the name

        if isinstance(b, datetime):
            if dob is None:
                dob = b.strftime('%Y-%m-%d')
        elif isinstance(b, str):
            bv = b.strip()
            if re.match(r'^MD\d+$', bv, re.IGNORECASE) and md_id is None:
                md_id = bv.upper()

    return last_name, first_name, f_id, assessment_date, md_id, dob


def _normalize_dob(dob: str) -> str:
    """Return the alternate DOB format: YYYY-MM-DD <-> MM-DD-YYYY."""
    try:
        if re.match(r'^\d{4}-\d{2}-\d{2}$', dob):
            p = dob.split('-')
            return f"{p[1]}-{p[2]}-{p[0]}"
        if re.match(r'^\d{2}-\d{2}-\d{4}$', dob):
            p = dob.split('-')
            return f"{p[2]}-{p[0]}-{p[1]}"
    except Exception:
        pass
    return dob


def _resolve_customer(cur, last_name: str, first_name: str,
                      md_id: Optional[str], dob: Optional[str]):
    """
    Find an existing customer.  Returns (customer_id, customer_code, found_without_md).

    Lookup order:
      1. MD ID  (authoritative — never reject on name mismatch)
      2. Name + DOB  (checks both date-format variants)
      3. Name only   (case-insensitive)
    """
    last_u  = last_name.upper().strip()
    first_u = first_name.upper().strip()
    md_norm = md_id.upper().strip() if md_id else None

    if md_norm:
        cur.execute(
            "SELECT id, customer_code FROM customers "
            "WHERE UPPER(TRIM(id_number)) = ?",
            (md_norm,)
        )
        row = cur.fetchone()
        if row:
            return row['id'], row['customer_code'], False

    if dob:
        dob_alt = _normalize_dob(dob)
        cur.execute(
            "SELECT id, customer_code FROM customers "
            "WHERE UPPER(TRIM(last_name)) = ? AND UPPER(TRIM(first_name)) = ? "
            "AND (date_of_birth = ? OR date_of_birth = ?)",
            (last_u, first_u, dob, dob_alt)
        )
        row = cur.fetchone()
        if row:
            return row['id'], row['customer_code'], True

    cur.execute(
        "SELECT id, customer_code FROM customers "
        "WHERE UPPER(TRIM(last_name)) = ? AND UPPER(TRIM(first_name)) = ?",
        (last_u, first_u)
    )
    row = cur.fetchone()
    if row:
        return row['id'], row['customer_code'], True

    return None, None, False


# ---------------------------------------------------------------------------
# Service entry collection (all rows in group, cleaned + merged)
# ---------------------------------------------------------------------------

def _collect_service_entries(
    group_rows: List[int], ws, year: int, rates: Dict[str, float]
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Returns (valid_entries, invalid_entries).
    invalid_entries includes any non-empty, non-numeric string that couldn't be parsed —
    e.g. "Accepted", "Approved", "3654/12122/2025".
    Plain numbers (auth unit counts) are silently ignored.

    `rates` (service name -> rate_per_day) is the live lookup from the
    `services` table — the caller must have already verified every mapped
    service has a rate (see _require_rates_for_mapped_services).
    """
    entries: List[Dict[str, Any]] = []
    invalid_entries: List[Dict[str, Any]] = []

    for col_idx, (svc_code, col_type) in SERVICE_COLS.items():
        rate = rates[svc_code]
        raw_vals = [
            ws.cell(r, col_idx).value
            for r in group_rows
            if ws.cell(r, col_idx).value is not None
        ]
        if not raw_vals:
            continue

        if col_type == 'units':
            for v in raw_vals:
                parsed, parse_err = _parse_h0038_cell(v, year, rate)
                if parsed:
                    for entry in parsed:
                        entry['service_name'] = svc_code
                        entries.append(entry)
                elif not isinstance(v, (int, float)):
                    sv = str(v).strip()
                    if sv:
                        invalid_entries.append({
                            'service_name': svc_code,
                            'raw_value': sv,
                            'error': parse_err or 'Unrecognized value — expected M/D-N (e.g. 3/5-2)',
                        })
        else:
            all_dates: List[date] = []
            for v in raw_vals:
                parsed_dates, parse_err = _parse_date_cell(v, year)
                if parsed_dates:
                    all_dates.extend(parsed_dates)
                elif not isinstance(v, (int, float)):
                    sv = str(v).strip()
                    if sv:
                        invalid_entries.append({
                            'service_name': svc_code,
                            'raw_value': sv,
                            'error': parse_err or 'Unrecognized value — expected a date (M/D or M/D/YYYY)',
                        })
            for entry in _dates_to_range_entries(all_dates, rate):
                entry['service_name'] = svc_code
                entries.append(entry)

    return entries, invalid_entries


# ---------------------------------------------------------------------------
# Import endpoint
# ---------------------------------------------------------------------------

@router.post("/import", status_code=status.HTTP_201_CREATED)
def import_billing(
    file: UploadFile = File(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
):
    """
    Import billing data from an Excel file (.xlsx/.xls) — IOP multi-row format.

    Customer layout (cols 1-based):
      Header row : A=seq#  B=LAST  C=FIRST  D=F-diagnosis(F ID)  F=assessment  G-Q=services
      Sub-rows   : B=MD ID ('MD...') | DOB (datetime) | address

    Service columns (mapped):
      G(7)=H0005  H(8)=H0004  J(10)=H0015  L(12)=H2036
      M(13)=UA  O(15)=H0024  P(16)=H0038 (units M/D-N, 15min per unit)
    Unmapped (data ignored):
      I(9)=IOP AUTH UNITS  K(11)=PHP AUTH UNITS  N(14)=Peer Group Units  Q(17)=Peer Individual Session

    Cleaning applied before inserting:
      - Year added to bare M/D dates/ranges
      - Consecutive dates merged into single range entries
      - Noise removed: Approved, PENDING, ?????, auth-unit numbers
    """
    ext = os.path.splitext(file.filename)[-1].lower()
    if ext not in ['.xlsx', '.xls']:
        raise HTTPException(status_code=400,
                            detail="Unsupported file type. Please upload .xlsx or .xls.")
    try:
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            wb = openpyxl.load_workbook(file.file, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400,
                            detail=f"Failed to parse Excel file: {str(e)}")

    return _process_workbook(wb, file.filename)


def _process_workbook(wb, filename: str, source_type: str = 'excel', source_url: Optional[str] = None) -> Dict[str, Any]:
    """Run the IOP import over an already-loaded workbook and persist results.

    `source_type` is 'excel' for an uploaded file or 'google' for a Sheets/Drive
    sync — surfaced as the File Type column in import history.

    Shared by the file-upload endpoint and the Google Sheets/Drive sync endpoint.
    Because service entries are de-duplicated on (customer, service, start_date),
    re-running this over an updated sheet only adds newly-changed rows.
    """
    conn = get_db_connection()
    cur  = conn.cursor()
    total_customers = 0
    total_entries   = 0
    errors          = []
    customer_logs   = []
    rates           = _get_service_rates(cur)
    try:
        _require_rates_for_mapped_services(rates)
    except HTTPException:
        conn.close()
        raise

    iop_sheets = [s for s in wb.sheetnames if 'IOP' in s.upper()]
    if not iop_sheets:
        raise HTTPException(
            status_code=400,
            detail=f"No sheet containing 'IOP' found. Available sheets: {', '.join(wb.sheetnames)}"
        )

    for sheet_name in iop_sheets:
        ws = wb[sheet_name]

        for start_row, group_rows in _find_member_blocks(ws):
            last_name = first_name = ''

            try:
                (last_name, first_name, f_id,
                 assessment_date, md_id, dob) = _extract_customer_info(
                    start_row, group_rows, ws
                )

                if not last_name:
                    reason = (
                        f"Skipping member block (rows {group_rows[0]}-{group_rows[-1]}, "
                        f"sheet '{sheet_name}'): no customer name found anywhere in the block. "
                        f"Detected — memberId: {md_id or 'none'}, dob: {dob or 'none'}."
                    )
                    logger.warning(reason)
                    customer_logs.append({
                        'customer_name':   f"Unidentified member (rows {group_rows[0]}-{group_rows[-1]})",
                        'customer_id':     None,
                        'is_new_customer': False,
                        'is_error':        True,
                        'error_message':   reason,
                    })
                    continue

                year                          = _infer_year(group_rows, ws)
                service_entries, invalid_entries = _collect_service_entries(group_rows, ws, year, rates)

                if not service_entries and not invalid_entries:
                    continue

                created_at = datetime.utcnow().isoformat()

                # ----------------------------------------------------------
                # Resolve customer by ID first — no duplicate creation
                # ----------------------------------------------------------
                customer_id, customer_code, found_without_md = _resolve_customer(
                    cur, last_name, first_name, md_id, dob
                )
                is_new = customer_id is None

                if customer_id is not None:
                    # Back-fill MD ID so future imports always hit the ID lookup
                    if found_without_md and md_id:
                        cur.execute(
                            "UPDATE customers SET id_number = ? "
                            "WHERE id = ? AND (id_number IS NULL OR TRIM(id_number) = '')",
                            (md_id.upper().strip(), customer_id)
                        )
                else:
                    # Genuinely new customer
                    rand_str      = ''.join(random.choices(
                        string.ascii_uppercase + string.digits, k=6))
                    customer_code = f"{last_name}_{first_name}_{created_at[:10]}_{rand_str}"
                    cur.execute(
                        "INSERT INTO customers "
                        "(customer_code, last_name, first_name, date_of_birth, "
                        " active_status, id_number, f_id_number, created_at) "
                        "VALUES (?,?,?,?,?,?,?,?) RETURNING id",
                        (customer_code, last_name, first_name, dob,
                         'active', md_id, f_id, created_at)
                    )
                    customer_id = cur.fetchone()['id']
                    total_customers += 1

                # ----------------------------------------------------------
                # Insert service entries — skip any (service, start_date) already present
                # ----------------------------------------------------------
                cur.execute(
                    "SELECT service_name, start_date FROM customer_entries "
                    "WHERE customer_id = ?",
                    (customer_id,)
                )
                existing = {(r['service_name'], r['start_date']) for r in cur.fetchall()}

                batch_id = customer_code   # stable per customer — keeps one row in frontend
                entries_added: List[Dict[str, Any]] = []
                entries_skipped_count = 0
                for entry in service_entries:
                    key = (entry['service_name'], entry['start_date'])
                    if key in existing:
                        entries_skipped_count += 1
                        continue
                    cur.execute(
                        "INSERT INTO customer_entries "
                        "(customer_id, customer_code, service_name, start_date, end_date, "
                        " days, units, rate_per_day, amount_billed, amount_paid, created_at, batch_id) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                        (
                            customer_id,           customer_code,
                            entry['service_name'],
                            entry['start_date'],   entry['end_date'],
                            entry['days'],         entry.get('units'),
                            entry['rate_per_day'],
                            entry['amount_billed'], 0.0,
                            created_at,            batch_id,
                        )
                    )
                    existing.add(key)
                    total_entries += 1
                    entries_added.append({
                        'service_name': entry['service_name'],
                        'start_date':   entry['start_date'],
                        'end_date':     entry['end_date'],
                        'days':         entry['days'],
                        'units':        entry.get('units'),
                        'rate_per_day': entry['rate_per_day'],
                        'amount_billed': entry['amount_billed'],
                    })

                conn.commit()
                customer_logs.append({
                    'customer_name':    f"{last_name}, {first_name}",
                    'customer_id':      customer_id,
                    'is_new_customer':  is_new,
                    'entries_added':    entries_added,
                    'entries_skipped':  entries_skipped_count,
                    'entries_invalid':  invalid_entries,
                })

            except Exception as exc:
                conn.rollback()
                import traceback; traceback.print_exc()
                errors.append({
                    'sheet':    sheet_name,
                    'row':      start_row,
                    'customer': f"{last_name}, {first_name}",
                    'error':    str(exc),
                })
                customer_logs.append({
                    'customer_name':   f"{last_name}, {first_name}" if last_name else f"Row {start_row}",
                    'customer_id':     None,
                    'is_new_customer': False,
                    'is_error':        True,
                    'error_message':   str(exc),
                    'entries_added':   [],
                    'entries_skipped': 0,
                })

    # Persist this import session to import_logs
    log_id = None
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS import_logs (
                id SERIAL PRIMARY KEY,
                filename TEXT NOT NULL,
                imported_at TIMESTAMP DEFAULT NOW(),
                entries_added INTEGER DEFAULT 0,
                entries_skipped INTEGER DEFAULT 0,
                total_billed NUMERIC(12,2) DEFAULT 0,
                customers_new INTEGER DEFAULT 0,
                customers_existing INTEGER DEFAULT 0,
                customer_logs TEXT DEFAULT '[]',
                error_count INTEGER DEFAULT 0,
                entries_invalid INTEGER DEFAULT 0,
                source_type TEXT DEFAULT 'excel'
            )
        """)
        cur.execute(
            "ALTER TABLE import_logs ADD COLUMN IF NOT EXISTS entries_invalid INTEGER DEFAULT 0"
        )
        cur.execute(
            "ALTER TABLE import_logs ADD COLUMN IF NOT EXISTS source_type TEXT DEFAULT 'excel'"
        )
        cur.execute(
            "ALTER TABLE import_logs ADD COLUMN IF NOT EXISTS source_url TEXT"
        )
        total_billed_sum = round(sum(
            sum(e.get('amount_billed', 0) for e in log.get('entries_added', []))
            for log in customer_logs
        ), 2)
        total_skipped = sum(log.get('entries_skipped', 0) for log in customer_logs)
        total_invalid = sum(len(log.get('entries_invalid') or []) for log in customer_logs)
        customers_new_count = sum(1 for log in customer_logs if log.get('is_new_customer'))
        customers_existing_count = sum(
            1 for log in customer_logs
            if not log.get('is_new_customer') and not log.get('is_error')
        )
        cur.execute(
            """INSERT INTO import_logs
               (filename, entries_added, entries_skipped, total_billed,
                customers_new, customers_existing, customer_logs, error_count,
                entries_invalid, source_type, source_url)
               VALUES (?,?,?,?,?,?,?,?,?,?,?) RETURNING id""",
            (
                filename, total_entries, total_skipped, total_billed_sum,
                customers_new_count, customers_existing_count,
                json.dumps(customer_logs), len(errors), total_invalid,
                source_type, source_url,
            )
        )
        log_id = cur.fetchone()['id']
        conn.commit()
    except Exception as log_exc:
        print(f"WARNING: Failed to save import log: {log_exc}")
        try:
            conn.rollback()
        except Exception:
            pass

    conn.close()

    total_invalid = sum(len(log.get('entries_invalid') or []) for log in customer_logs)
    msg = f'Imported {total_customers} new customers and {total_entries} service entries.'
    if total_invalid:
        msg += f' {total_invalid} cell(s) had invalid date/unit format.'

    return {
        'message':              msg,
        'sheets_processed':     len(iop_sheets),
        'customers_inserted':   total_customers,
        'entries_inserted':     total_entries,
        'entries_invalid_count': total_invalid,
        'errors':               errors,
        'log_id':               log_id,
    }


def _ensure_import_logs_table(cur, conn):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS import_logs (
            id SERIAL PRIMARY KEY,
            filename TEXT NOT NULL,
            imported_at TIMESTAMP DEFAULT NOW(),
            entries_added INTEGER DEFAULT 0,
            entries_skipped INTEGER DEFAULT 0,
            total_billed NUMERIC(12,2) DEFAULT 0,
            customers_new INTEGER DEFAULT 0,
            customers_existing INTEGER DEFAULT 0,
            customer_logs TEXT DEFAULT '[]',
            error_count INTEGER DEFAULT 0,
            entries_invalid INTEGER DEFAULT 0,
            source_type TEXT DEFAULT 'excel'
        )
    """)
    cur.execute(
        "ALTER TABLE import_logs ADD COLUMN IF NOT EXISTS entries_invalid INTEGER DEFAULT 0"
    )
    cur.execute(
        "ALTER TABLE import_logs ADD COLUMN IF NOT EXISTS source_type TEXT DEFAULT 'excel'"
    )
    cur.execute(
        "ALTER TABLE import_logs ADD COLUMN IF NOT EXISTS source_url TEXT"
    )
    conn.commit()


@router.get("/import/logs")
def get_import_logs(
    limit: int = Query(100, ge=1, le=500),
    credentials: HTTPAuthorizationCredentials = Depends(security),
):
    """Return recent import sessions in reverse-chronological order."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_import_logs_table(cur, conn)
        cur.execute(
            """SELECT id, filename, imported_at, entries_added, entries_skipped,
                      total_billed, customers_new, customers_existing, customer_logs, error_count,
                      COALESCE(entries_invalid, 0) AS entries_invalid,
                      COALESCE(source_type, 'excel') AS source_type,
                      source_url
               FROM import_logs
               ORDER BY imported_at DESC
               LIMIT ?""",
            (limit,)
        )
        rows = cur.fetchall()
        result = []
        for row in rows:
            r = dict(row)
            cl = r.get('customer_logs')
            if isinstance(cl, str):
                try:
                    r['customer_logs'] = json.loads(cl)
                except Exception:
                    r['customer_logs'] = []
            elif cl is None:
                r['customer_logs'] = []
            # imported_at is a naive TIMESTAMP written by NOW() (UTC on the DB
            # server). Tag it as UTC so the browser converts to EST correctly
            # instead of treating it as local time.
            ia = r.get('imported_at')
            if hasattr(ia, 'isoformat'):
                if ia.tzinfo is None:
                    ia = ia.replace(tzinfo=timezone.utc)
                r['imported_at'] = ia.isoformat()
            # Cast Decimal to float for JSON serialisation
            if r.get('total_billed') is not None:
                r['total_billed'] = float(r['total_billed'])
            result.append(r)
        return result
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Fix invalid import cells from history (user edits in UI)
# ---------------------------------------------------------------------------

class InvalidFixItem(BaseModel):
    service_name: str
    raw_value: str = Field(..., min_length=1)
    original_raw_value: Optional[str] = None


class FixInvalidPayload(BaseModel):
    customer_id: int
    fixes: List[InvalidFixItem]
    year: Optional[int] = None


def _parse_raw_to_entries(
    service_name: str, raw_value: str, year: int, rates: Dict[str, float]
) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """Parse a corrected cell string into DB-ready entry dicts.

    `rates` is the live rate_per_day lookup from the `services` table — the
    only source of truth for billing rates.
    """
    col_type = SERVICE_COL_TYPES.get(service_name)
    if not col_type:
        return [], f'Unknown service: {service_name}'
    rate = rates.get(service_name)
    if rate is None:
        return [], f"No rate_per_day configured for '{service_name}' — set it on the Services page first."
    sv = str(raw_value).strip()
    if not sv:
        return [], 'Value is required'

    if col_type == 'units':
        parsed, err = _parse_h0038_cell(sv, year, rate)
        if parsed:
            for e in parsed:
                e['service_name'] = service_name
            return parsed, None
        return [], err

    all_dates: List[date] = []
    for token in re.split(r'[,;]', sv):
        dates, err = _parse_one_date_token(token.strip(), year)
        if dates:
            all_dates.extend(dates)
        elif err:
            return [], err
    if not all_dates:
        return [], 'Invalid date format — expected M/D or M/D/YYYY'
    entries = _dates_to_range_entries(all_dates, rate)
    for e in entries:
        e['service_name'] = service_name
    return entries, None


def _insert_parsed_entries(
    cur, customer_id: int, customer_code: str, service_entries: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], int]:
    """Insert parsed entries; skip duplicates. Returns (added, skipped_count)."""
    cur.execute(
        "SELECT service_name, start_date FROM customer_entries WHERE customer_id = ?",
        (customer_id,),
    )
    existing = {(r['service_name'], r['start_date']) for r in cur.fetchall()}
    created_at = datetime.utcnow().isoformat()
    batch_id = customer_code
    added: List[Dict[str, Any]] = []
    skipped = 0
    for entry in service_entries:
        key = (entry['service_name'], entry['start_date'])
        if key in existing:
            skipped += 1
            continue
        cur.execute(
            "INSERT INTO customer_entries "
            "(customer_id, customer_code, service_name, start_date, end_date, "
            " days, units, rate_per_day, amount_billed, amount_paid, created_at, batch_id) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                customer_id, customer_code,
                entry['service_name'],
                entry['start_date'], entry['end_date'],
                entry['days'], entry.get('units'),
                entry['rate_per_day'],
                entry['amount_billed'], 0.0,
                created_at, batch_id,
            ),
        )
        existing.add(key)
        added.append({
            'service_name': entry['service_name'],
            'start_date': entry['start_date'],
            'end_date': entry['end_date'],
            'days': entry['days'],
            'units': entry.get('units'),
            'rate_per_day': entry['rate_per_day'],
            'amount_billed': entry['amount_billed'],
        })
    return added, skipped


@router.post("/import/logs/{log_id}/fix-invalid")
def fix_invalid_import_entries(
    log_id: int,
    payload: FixInvalidPayload = Body(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
):
    """
    Re-parse corrected raw values from import history and insert valid service entries.
    Removes resolved items from entries_invalid on the import log.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_import_logs_table(cur, conn)
        cur.execute(
            "SELECT id, customer_logs, entries_added, entries_skipped, "
            "total_billed, entries_invalid FROM import_logs WHERE id = ?",
            (log_id,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail='Import log not found')

        logs = row['customer_logs']
        if isinstance(logs, str):
            try:
                customer_logs = json.loads(logs)
            except Exception:
                customer_logs = []
        else:
            customer_logs = logs or []

        cust_log = next(
            (l for l in customer_logs if l.get('customer_id') == payload.customer_id),
            None,
        )
        if not cust_log:
            raise HTTPException(status_code=404, detail='Customer not found in this import')

        cur.execute(
            'SELECT id, customer_code FROM customers WHERE id = ?',
            (payload.customer_id,),
        )
        cust_row = cur.fetchone()
        if not cust_row:
            raise HTTPException(status_code=404, detail='Customer not found')

        year = payload.year or datetime.utcnow().year
        rates = _get_service_rates(cur)
        still_invalid: List[Dict[str, Any]] = []
        parse_errors: List[Dict[str, Any]] = []
        resolved_keys = set()

        all_parsed: List[Dict[str, Any]] = []
        for fix in payload.fixes:
            orig = (fix.original_raw_value or fix.raw_value).strip()
            resolved_keys.add((fix.service_name, orig))
            parsed, err = _parse_raw_to_entries(fix.service_name, fix.raw_value, year, rates)
            if err:
                parse_errors.append({
                    'service_name': fix.service_name,
                    'raw_value': fix.raw_value,
                    'error': err,
                })
                still_invalid.append({
                    'service_name': fix.service_name,
                    'raw_value': fix.raw_value,
                    'error': err,
                })
            else:
                all_parsed.extend(parsed)

        added, skipped = _insert_parsed_entries(
            cur, payload.customer_id, cust_row['customer_code'], all_parsed
        )

        remaining_invalid = [
            inv for inv in (cust_log.get('entries_invalid') or [])
            if (inv.get('service_name'), (inv.get('raw_value') or '').strip()) not in resolved_keys
        ]
        remaining_invalid.extend(still_invalid)

        cust_log['entries_invalid'] = remaining_invalid
        cust_log['entries_added'] = (cust_log.get('entries_added') or []) + added
        cust_log['entries_skipped'] = (cust_log.get('entries_skipped') or 0) + skipped

        total_billed = round(
            sum(
                sum(e.get('amount_billed', 0) for e in log.get('entries_added', []))
                for log in customer_logs
            ),
            2,
        )
        total_added = sum(len(log.get('entries_added') or []) for log in customer_logs)
        total_skipped = sum(log.get('entries_skipped', 0) for log in customer_logs)
        total_invalid = sum(len(log.get('entries_invalid') or []) for log in customer_logs)

        cur.execute(
            """UPDATE import_logs
               SET customer_logs = ?, entries_added = ?, entries_skipped = ?,
                   total_billed = ?, entries_invalid = ?
               WHERE id = ?""",
            (
                json.dumps(customer_logs),
                total_added,
                total_skipped,
                total_billed,
                total_invalid,
                log_id,
            ),
        )
        conn.commit()

        return {
            'entries_added': len(added),
            'entries_skipped': skipped,
            'entries_invalid_remaining': len(remaining_invalid),
            'parse_errors': parse_errors,
            'added': added,
        }
    except HTTPException:
        raise
    except Exception as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Google Sheets / Google Drive sync
# ---------------------------------------------------------------------------
# A public Google Sheet or Drive-hosted .xlsx can be fetched directly:
#   Sheets  https://docs.google.com/spreadsheets/d/<ID>/edit#gid=0
#           -> https://docs.google.com/spreadsheets/d/<ID>/export?format=xlsx
#   Drive   https://drive.google.com/file/d/<ID>/view
#           -> https://drive.google.com/uc?export=download&id=<ID>
# The sheet/file must be shared as "Anyone with the link" for this to work.
# ---------------------------------------------------------------------------

_DL_TIMEOUT = 30            # seconds
_MAX_DL_BYTES = 25 * 1024 * 1024   # 25 MB safety cap


def _google_download_url(url: str) -> str:
    """Convert a Google Sheets/Drive share link into a direct .xlsx download URL.

    Passes through non-Google URLs unchanged (must point at a downloadable file).
    Raises HTTPException(400) if a Google link is recognised but has no file id.
    """
    u = (url or '').strip()
    if not u:
        raise HTTPException(status_code=400, detail='Sync URL is empty.')

    # Google Sheets — export the whole workbook as xlsx
    m = re.search(r'docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)', u)
    if m:
        return f'https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx'

    # Google Drive file — /file/d/<ID>/ or ?id=<ID> or open?id=<ID>
    if 'drive.google.com' in u or 'docs.google.com' in u:
        m = re.search(r'/file/d/([a-zA-Z0-9_-]+)', u)
        if not m:
            m = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', u)
        if m:
            return f'https://drive.google.com/uc?export=download&id={m.group(1)}'
        raise HTTPException(
            status_code=400,
            detail='Could not find a file id in that Google link. '
                   'Use a Google Sheets link or a Drive "share" link.'
        )

    # Not a Google URL — use as-is (must return xlsx bytes)
    return u


def _filename_from_response(resp, fallback: str) -> str:
    """Pull the real sheet/file name out of the Content-Disposition header.

    Google returns e.g.  attachment; filename="Billing Sheet.xlsx"
    or the RFC 5987 form  filename*=UTF-8''Billing%20Sheet.xlsx
    """
    cd = resp.headers.get('Content-Disposition') or ''
    if cd:
        # RFC 5987 form takes precedence — it carries the correct encoding
        m = re.search(r"filename\*\s*=\s*[^']*'[^']*'([^;]+)", cd, re.IGNORECASE)
        if m:
            try:
                name = unquote(m.group(1).strip().strip('"'))
                if name:
                    return name
            except Exception:
                pass
        m = re.search(r'filename\s*=\s*"([^"]+)"', cd, re.IGNORECASE)
        if not m:
            m = re.search(r'filename\s*=\s*([^;]+)', cd, re.IGNORECASE)
        if m:
            name = m.group(1).strip().strip('"')
            if name:
                return name
    return fallback


def _fetch_workbook_from_url(url: str):
    """Download the sheet/file at `url`.

    Returns (workbook, filename) where filename is the real Google Sheets /
    Drive document name when the response advertises one.
    """
    dl_url = _google_download_url(url)
    try:
        resp = requests.get(dl_url, timeout=_DL_TIMEOUT, allow_redirects=True)
    except requests.RequestException as e:
        raise HTTPException(status_code=502,
                            detail=f'Failed to reach the sync URL: {e}')

    if resp.status_code != 200:
        raise HTTPException(
            status_code=502,
            detail=f'Sync URL returned HTTP {resp.status_code}. '
                   'Make sure the sheet is shared as "Anyone with the link".'
        )

    content = resp.content
    if len(content) > _MAX_DL_BYTES:
        raise HTTPException(status_code=413, detail='Sync file is too large.')

    # A shared-but-restricted sheet returns an HTML sign-in page, not xlsx.
    ctype = (resp.headers.get('Content-Type') or '').lower()
    if content[:15].lstrip().lower().startswith(b'<!doctype html') or 'text/html' in ctype:
        raise HTTPException(
            status_code=400,
            detail='The link did not return a spreadsheet. Confirm sharing is set '
                   'to "Anyone with the link" (Viewer) and the link is correct.'
        )

    try:
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400,
                            detail=f'Failed to parse the synced spreadsheet: {e}')

    fallback = f'Google Sync {datetime.utcnow().strftime("%Y-%m-%d %H:%M")}'
    return wb, _filename_from_response(resp, fallback)


class SyncNowPayload(BaseModel):
    sync_url: str = Field(..., min_length=1)


@router.post("/sync", status_code=status.HTTP_201_CREATED)
def sync_from_url(
    payload: SyncNowPayload = Body(...),
    credentials: HTTPAuthorizationCredentials = Depends(security),
):
    """Download the linked Google Sheet/Drive file and import it.

    The URL is never persisted server-side — the caller must pass it on every
    sync (the frontend keeps it in memory only, per-session, and clears it
    when the user clicks Clear). Re-running only adds newly-changed rows.
    """
    url = payload.sync_url.strip()
    if not url:
        raise HTTPException(
            status_code=400,
            detail='No sync URL provided. Paste a Google Sheets/Drive link first.'
        )
    # Validate it resolves to a downloadable link before attempting the download.
    _google_download_url(url)

    wb, filename = _fetch_workbook_from_url(url)
    result = _process_workbook(wb, filename, source_type='google', source_url=url)
    result['synced_from'] = url
    result['filename'] = filename
    return result
